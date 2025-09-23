import os, glob
from datetime import datetime
from flask import Flask, render_template, request, send_from_directory, redirect, url_for, flash
from flask_cors import CORS
import pandas as pd
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "change-me-to-a-random-string"

# ---------- Base folders ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config["UPLOAD_FOLDER"]  = os.path.join(BASE_DIR, "uploads")
app.config["RESULTS_FOLDER"] = os.path.join(BASE_DIR, "results")
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["RESULTS_FOLDER"], exist_ok=True)

# ---------- Security: server access key ----------
# Set ACCESS_KEY on Render (Environment tab). If empty, the check is disabled (handy for local dev).
ACCESS_KEY = os.environ.get("ACCESS_KEY", "").strip()

def require_access_key() -> bool:
    if not ACCESS_KEY:
        return True  # allow locally if no key configured
    provided = (request.form.get("access_key") or "").strip()
    return provided == ACCESS_KEY

# ---------- Template resolver (pick newest .xlsm in /template) ----------
TEMPLATE_DIR = os.path.join(BASE_DIR, "template")

def resolve_template_path() -> str:
    pattern = os.path.join(TEMPLATE_DIR, "*.xlsm")
    matches = glob.glob(pattern)
    if not matches:
        return ""
    return max(matches, key=lambda p: os.path.getmtime(p))

# 5 MB upload cap
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024

# CORS: restrict to your domains
CORS(app, resources={r"/*": {"origins": [
    "https://www.easynetballtrials.com.au",
    "https://editor.wix.com",
    "https://easy-netball-trials-app.onrender.com"
]}})

# Optional: silence favicon 404s
@app.route("/favicon.ico")
def favicon():
    return ("", 204)

ALLOWED_EXT = {".csv"}
POSITIONS = ["GS", "GA", "WA", "C", "WD", "GD", "GK"]

# ---------- CSV loading & normalisation ----------
REQUIRED = ["First Name", "Last Name", "PreferredPos1", "PreferredPos2", "PreferredPos3"]
OPTIONAL = ["Active", "Seed"]

def normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {}
    for c in df.columns:
        cc = c.strip()
        up = cc.replace(" ", "").lower()
        if up in ("firstname","first_name","first"):
            mapping[c] = "First Name"
        elif up in ("lastname","last_name","surname","last"):
            mapping[c] = "Last Name"
        elif up in ("preferredpos1","preferredposition1","pos1","preferredposi","preferredpos"):
            mapping[c] = "PreferredPos1"
        elif up in ("preferredpos2","preferredposition2","pos2"):
            mapping[c] = "PreferredPos2"
        elif up in ("preferredpos3","preferredposition3","pos3"):
            mapping[c] = "PreferredPos3"
        elif up in ("active","attendance","attend","present"):
            mapping[c] = "Active"
        elif up in ("seed","seeding","rank"):
            mapping[c] = "Seed"
        else:
            mapping[c] = cc
    df = df.rename(columns=mapping)
    for col in REQUIRED:
        if col not in df.columns:
            df[col] = ""
    for col in OPTIONAL:
        if col not in df.columns:
            df[col] = ""
    return df

def load_players(file_storage) -> pd.DataFrame:
    df = pd.read_csv(file_storage)
    df = normalize_headers(df)
    # defaults
    df["Active"] = df["Active"].fillna("").astype(str).str.strip().str.upper()
    df.loc[df["Active"] == "", "Active"] = "YES"
    df["Seed"] = pd.to_numeric(df["Seed"], errors="coerce").fillna(3).astype(int)
    for p in ["PreferredPos1","PreferredPos2","PreferredPos3"]:
        df[p] = df[p].fillna("").astype(str).str.strip().str.upper()
    df["Full Name"] = (df["First Name"].fillna("").astype(str).str.strip() + " " +
                       df["Last Name"].fillna("").astype(str).str.strip()).str.strip()
    return df

# ---------- Scheduling core ----------
def capacity_ok(num_active, num_courts, rounds, min_games=3):
    capacity = num_courts * 14 * rounds  # 7 pos * 2 teams * courts * rounds
    return capacity >= num_active * min_games, capacity

def pref_tag_for(full: str, pos: str, prefs_by_full: dict) -> str | None:
    rec = prefs_by_full.get(full, {})
    pos = (pos or "").upper()
    if not rec:
        return None
    if pos == rec.get("p1"):
        return "P1"
    if pos == rec.get("p2"):
        return "P2"
    if pos == rec.get("p3"):
        return "P3"
    return None

def pick_next(pool, pos, games_played, court_hist, used_this_round,
              min_games_target, prefer_hist, prefs_by_full):
    pos = pos.upper()
    pref_weight = {"P1": 40, "P2": 25, "P3": 8}
    target_share = {"P1": 0.45, "P2": 0.35, "P3": 0.20}  # soft target mix

    for pass_num in (1, 2):
        best_score = -10**9
        best = None
        for p in pool:
            full = p["full"]
            if full in used_this_round:
                continue
            if pos not in (p["p1"], p["p2"], p["p3"]):
                continue
            gp = int(games_played.get(full, 0))
            if pass_num == 1 and gp >= min_games_target:
                continue

            score = 1000 - gp * 20 - int(p["seed"]) * 2

            # Court-spread bonus
            if court_hist.get(full, "").find(str(p["court"])) == -1:
                score += 10

            # Preference-aware scoring
            tag = pref_tag_for(full, pos, prefs_by_full)
            if tag:
                count_same = prefer_hist[full].get(tag, 0)
                score += pref_weight.get(tag, 0) - 12 * count_same
                if gp > 0:
                    want = target_share[tag] * gp
                    diff = want - count_same
                    score += 8 * diff

            if score > best_score:
                best_score = score
                best = full

        if best:
            return best
    return ""

def schedule_rounds(df_active, rounds, courts, min_games_target=3):
    games_played = {row["Full Name"]: 0 for _, row in df_active.iterrows()}
    court_hist    = {row["Full Name"]: "" for _, row in df_active.iterrows()}

    pool_base, prefs_by_full = [], {}
    for _, row in df_active.iterrows():
        full = row["Full Name"]
        p1 = row["PreferredPos1"]; p2 = row["PreferredPos2"]; p3 = row["PreferredPos3"]
        pool_base.append({"full": full, "p1": p1, "p2": p2, "p3": p3, "seed": int(row["Seed"])})
        prefs_by_full[full] = {"p1": p1, "p2": p2, "p3": p3}

    prefer_hist = {full: {"P1": 0, "P2": 0, "P3": 0} for full in prefs_by_full.keys()}

    def bump_hist(full: str, pos: str):
        tag = pref_tag_for(full, pos, prefs_by_full)
        if tag:
            prefer_hist[full][tag] += 1

    rows = []
    for rnd in range(1, rounds + 1):
        used = set()
        for court in range(1, courts + 1):
            pool = [dict(p, court=court) for p in pool_base]
            teamA, teamB = [], []
            for pos in POSITIONS:
                a = pick_next(pool, pos, games_played, court_hist, used, min_games_target, prefer_hist, prefs_by_full)
                if a:
                    used.add(a); games_played[a] += 1; court_hist[a] += f"{court},"; bump_hist(a, pos)
                teamA.append(a or "")
                b = pick_next(pool, pos, games_played, court_hist, used, min_games_target, prefer_hist, prefs_by_full)
                if b:
                    used.add(b); games_played[b] += 1; court_hist[b] += f"{court},"; bump_hist(b, pos)
                teamB.append(b or "")
            for i, pos in enumerate(POSITIONS):
                rows.append((rnd, court, "A", pos, teamA[i]))
                rows.append((rnd, court, "B", pos, teamB[i]))

    sched = pd.DataFrame(rows, columns=["Round", "Court", "Team", "Position", "Player"])
    return sched, games_played

# ---------- Routes ----------
@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")

@app.route("/generate", methods=["POST"])
def generate():
    # Gate with access key (prevents direct public posts to Render)
    if not require_access_key():
        flash("Unauthorized request.")
        return redirect(url_for("index"))

    # 1) Validate upload
    f = request.files.get("file")
    if not f or os.path.splitext(f.filename)[1].lower() not in ALLOWED_EXT:
        flash("Please upload a .csv file.")
        return redirect(url_for("index"))

    # 2) Load & normalize players
    try:
        df = load_players(f)
    except Exception as e:
        flash(f"Could not read CSV: {e}")
        return redirect(url_for("index"))

    # 3) Read form options
    try:
        rounds = int(request.form.get("rounds", "10"))
    except:
        rounds = 10
    try:
        courts = int(request.form.get("courts", "2"))
    except:
        courts = 2

    number_players = (request.form.get("number_players", "NO").upper() == "YES")
    show_prefs     = (request.form.get("show_prefs", "NO").upper() == "YES")

    df_active = df[df["Active"].str.upper() == "YES"].copy()
    num_active = len(df_active)

    ok, cap = capacity_ok(num_active, courts, rounds, min_games=3)
    if not ok:
        flash(f"Warning: capacity {cap} is less than the {num_active*3} needed to guarantee 3 games each.")

    # 4) Build schedule in Python
    schedule_df, games_played = schedule_rounds(df_active, rounds, courts, min_games_target=3)

    # 5) Open the latest template (.xlsm) with VBA preserved
    template_path = resolve_template_path()
    if not template_path:
        flash("Server template not found in /template. Please upload your .xlsm.")
        return redirect(url_for("index"))

    wb = load_workbook(template_path, keep_vba=True)

    # 6) Write user choices into a hidden Control sheet (read by your VBA)
    try:
        ws_ctrl = wb["Control"]
    except KeyError:
        ws_ctrl = wb.create_sheet("Control")
    ws_ctrl["B4"].value = "YES" if number_players else "NO"   # Add numbers to players?
    ws_ctrl["B5"].value = "YES" if show_prefs else "NO"       # Show preferred positions on boards?
    ws_ctrl.sheet_state = "hidden"

    # 7) Fill Players sheet
    ws_players = wb["Players"]
    ws_players.delete_rows(2, ws_players.max_row)

    numbers_map = {}
    if number_players:
        tmp = df[df["Active"].str.upper() == "YES"].copy()
        tmp["LastLower"]  = tmp["Last Name"].astype(str).str.lower()
        tmp["FirstLower"] = tmp["First Name"].astype(str).str.lower()
        tmp = tmp.sort_values(["LastLower", "FirstLower"], kind="mergesort")
        tmp["Number"] = range(1, len(tmp) + 1)
        numbers_map = dict(zip(tmp["Full Name"], tmp["Number"]))

    headers = [cell.value for cell in ws_players[1]]
    has_number_col = any((str(h or "").strip().lower() == "number") for h in headers)
    if number_players and not has_number_col:
        ws_players.cell(row=1, column=len(headers) + 1, value="Number")

    header_map = {str(ws_players.cell(row=1, column=c).value).strip().lower(): c
                  for c in range(1, ws_players.max_column + 1)}

    for _, r in df.iterrows():
        ws_players.append([
            r["First Name"], r["Last Name"], r["Full Name"],
            r["Active"], int(r["Seed"]), r["PreferredPos1"], r["PreferredPos2"], r["PreferredPos3"]
        ])
        if number_players:
            num = numbers_map.get(r["Full Name"], "")
            ws_players.cell(
                row=ws_players.max_row,
                column=header_map.get("number", ws_players.max_column),
                value=num
            )

    # 8) Fill Schedule sheet
    ws_sched = wb["Schedule"]
    ws_sched.delete_rows(2, ws_sched.max_row)
    for _, row in schedule_df.iterrows():
        ws_sched.append([
            int(row["Round"]), int(row["Court"]),
            row["Team"], row["Position"], row["Player"] or ""
        ])

    # 9) Fill GameTally
    ws_gt = wb["GameTally"]
    ws_gt.delete_rows(2, ws_gt.max_row)
    for _, row in df_active.iterrows():
        full = row["Full Name"]
        ws_gt.append([row["First Name"], row["Last Name"], full, int(games_played.get(full, 0)), ""])

    # 10) Save to /generated and show result page
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"netball_courtboards_{stamp}.xlsm"
    out_dir  = os.path.join(app.root_path, "generated")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, filename)
    wb.save(out_path)

    return render_template(
        "result.html",
        filename=filename,
        timestamp=datetime.now().strftime("%Y-%m-%d %H:%M"),
        rounds=rounds, courts=courts,
        number_players="YES" if number_players else "NO",
        show_prefs="YES" if show_prefs else "NO",
        players=num_active
    )

@app.route("/download/<path:filename>")
def download_file(filename):
    directory = os.path.join(app.root_path, "generated")
    return send_from_directory(directory, filename, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)
