# scheduler/excel_writer.py
import io
from typing import Iterable, Dict, List
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Expected normalized row keys coming from schema.validate_and_normalize_csv:
# first_name, last_name, active, number, p1, p2, p3, seed

PLAYERS_SHEET = "Players"

COLMAP = {
    "first_name": "A",
    "last_name":  "B",
    "active":     "C",
    "number":     "D",
    "p1":         "E",
    "p2":         "F",
    "p3":         "G",
    "seed":       "H",
}

BIB_RANGES = ("K19", "K20", "K22", "K23")  # left untouched

def _clear_players_table(ws: Worksheet):
    # Keep headers in row 1. Clear rows 2..N for cols A..H.
    max_row = ws.max_row or 1
    ws.delete_rows(2, max_row)  # wipe everything under headers

def _unlock_bib_cells(ws: Worksheet):
    # Make sure bib cells are editable even if sheet is protected in template
    for addr in BIB_RANGES:
        c = ws[addr]
        c.protection = c.protection.copy(locked=False)

def _strip_all_comments(wb):
    """Remove legacy comments/VML so Excel doesn't show 'repair' message."""
    for ws in wb.worksheets:
        # openpyxl 3.1+: ws.comments is list; setting cell.comment=None removes it
        for c in list(getattr(ws, "comments", [])):
            try:
                c.parent.comment = None
            except Exception:
                pass
        # Also clear any residual container openpyxl tracks
        try:
            ws._comments = []
        except Exception:
            pass

def inject_players_csv(template_path: str,
                       rows: Iterable[Dict[str, str]],
                       out_stream: io.BytesIO,
                       shell_mode: bool = True) -> None:
    """
    Write players into Players sheet of an xlsm template, keep VBA/macros,
    and avoid corrupting drawings/comments.
    """
    wb = load_workbook(template_path, keep_vba=True, data_only=False)

    if PLAYERS_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Template missing '{PLAYERS_SHEET}' sheet")

    ws = wb[PLAYERS_SHEET]

    # Ensure bib cells remain editable
    _unlock_bib_cells(ws)

    # Clear existing player data (rows 2..N)
    _clear_players_table(ws)

    # Write rows, mapping strictly by key -> column
    out_row = 2
    for r in rows:
        # Defaults
        first = r.get("first_name", "").strip()
        last  = r.get("last_name", "").strip()
        active = r.get("active", "").strip().upper()
        if active == "":
            active = "YES"            # blank means active
        number = r.get("number", "")
        p1 = r.get("p1", "").upper()
        p2 = r.get("p2", "").upper()
        p3 = r.get("p3", "").upper()
        seed = r.get("seed", "")
        if str(seed).strip() == "":
            seed = 3

        ws[f"{COLMAP['first_name']}{out_row}"] = first
        ws[f"{COLMAP['last_name']}{out_row}"]  = last
        ws[f"{COLMAP['active']}{out_row}"]     = active
        ws[f"{COLMAP['number']}{out_row}"]     = number
        ws[f"{COLMAP['p1']}{out_row}"]         = p1
        ws[f"{COLMAP['p2']}{out_row}"]         = p2
        ws[f"{COLMAP['p3']}{out_row}"]         = p3
        ws[f"{COLMAP['seed']}{out_row}"]       = seed
        out_row += 1

    # Strip legacy comments so Excel won't "repair"
    _strip_all_comments(wb)

    # Save to the provided stream (keeps VBA)
    wb.save(out_stream)
