# webbridge.py
"""
Small helpers for diagnostics and (optionally) forcing Control flags
after we write the macro-enabled template. This is optional; template_writer
already sets these flags. Use when you want extra safety or logs.
"""

from openpyxl import load_workbook
from typing import Iterable


def set_control_flags(
    xlsm_path: str,
    num_rounds: int,
    num_courts: int,
    numbering: bool,
    show_prefs: bool,
    pending: bool = True,
) -> None:
    """
    Open the workbook and (re)apply Control flags.
    Safe to call after template_writer.write_into_template.
    """
    wb = load_workbook(xlsm_path, keep_vba=True)
    ws = wb["Control"] if "Control" in wb.sheetnames else wb.create_sheet("Control")

    ws["B2"].value = int(num_rounds)               # Rounds
    ws["B3"].value = int(num_courts)               # Courts
    ws["B4"].value = "YES" if numbering else "NO"  # Numbering switch
    ws["B6"].value = "YES" if show_prefs else "NO" # Show preferred positions
    ws["Z1"].value = "PENDING" if pending else "DONE"

    wb.save(xlsm_path)


def sanity_log_players_header(xlsm_path: str) -> str:
    """
    Return a short string showing the Players header row so you can see
    the exact column order Render produced.
    """
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    ws = wb["Players"] if "Players" in wb.sheetnames else None
    if not ws:
        return "Players sheet not found."

    headers: Iterable[str] = []
    col = 1
    while True:
        v = ws.cell(row=1, column=col).value
        if v is None:
            break
        headers.append(str(v))
        col += 1

    return "Players headers: [" + ", ".join(headers) + "]"
