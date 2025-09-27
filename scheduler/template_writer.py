from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PLAYERS_SHEET = "Players"
CONTROL_SHEET = "Control"
PLAYER_HEADERS = [
    "First Name","Last Name","Active","Number",
    "PreferredPos1","PreferredPos2","PreferredPos3","Seed"
]
SHELL_CLEAR_SHEETS = ["CourtBoards","Schedule","GameTally"]

def _clear_data_below_header(ws):
    maxr = ws.max_row
    if maxr > 1:
        ws.delete_rows(2, maxr-1)

def inject_players_csv(template_path: str, rows: list[dict], out_stream, shell_mode: bool=False):
    wb = load_workbook(template_path, keep_vba=True, data_only=False)

    # Hide Control
    if CONTROL_SHEET in wb.sheetnames:
        try:
            wb[CONTROL_SHEET].sheet_state = "veryHidden"
        except Exception:
            wb[CONTROL_SHEET].sheet_state = "hidden"

    # Ship a shell (clear non-Players rows)
    if shell_mode:
        for nm in SHELL_CLEAR_SHEETS:
            if nm in wb.sheetnames:
                _clear_data_below_header(wb[nm])

    ws = wb[PLAYERS_SHEET]

    # Map headers
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str):
            header_map[v.strip()] = c

    missing = [h for h in PLAYER_HEADERS if h not in header_map]
    if missing:
        raise ValueError(f"Template missing headers: {missing}")

    # Clear and write
    _clear_data_below_header(ws)
    r_out = 2
    for row in rows:
        for h in PLAYER_HEADERS:
            ws.cell(row=r_out, column=header_map[h], value=row.get(h, ""))
        r_out += 1

    # GameTally Notes column width = 30
    if "GameTally" in wb.sheetnames:
        gt = wb["GameTally"]
        notes_col = None
        for c in range(1, gt.max_column + 1):
            if str(gt.cell(row=1, column=c).value).strip().lower() == "notes":
                notes_col = c; break
        if notes_col:
            gt.column_dimensions[get_column_letter(notes_col)].width = 30

    wb.save(out_stream)
